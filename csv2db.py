import csv, utm
from datetime import datetime
from openpyxl import load_workbook
from models import Crime, ShotSpotterEvent

CRIME_CSV_PATHS = [
    'data/crime_incidents_2011_CSV_LONLAT.csv',
    'data/crime_incidents_2012_CSV_LONLAT.csv',
    'data/crime_incidents_2013_CSV_LONLAT.csv'
]

SHOTSPOTTER_EXCEL_PATH = 'data/shotspotter.xlsx'
SHOTSPOTTER_CSV_PATH = 'data/shotspotter.csv'
SS_DATE_FORMAT = '%m/%d/%y %H:%M'
DC_CRIME_DATE_FORMAT = '%m/%d/%Y %I:%M:%S %p'

for path in CRIME_CSV_PATHS:
    with open(path, 'r') as csv_file:
        reader = csv.DictReader(csv_file, delimiter=',')
        for crime in reader:
            crime['lat'] = float(crime['LAT'])
            crime['lon'] = float(crime['LON'])
            del crime['LAT']
            del crime['LON']
            crime['utm_lat'], crime['utm_lon'], _, _ = utm.from_latlon(
                crime['lat'], crime['lon']
            )
            for k in ['START_DATE', 'END_DATE', 'REPORTDATETIME']:
                if k in crime:
                    try:
                        crime[k] = datetime.strptime(crime[k], DC_CRIME_DATE_FORMAT)
                    except ValueError:
                        del crime[k]
            c = Crime(crime)
            c.m.save()

wb = load_workbook(SHOTSPOTTER_EXCEL_PATH)
sh = wb.get_sheet_by_name(wb.sheetnames[0])
rows = [[cell.value for cell in row] for row in sh.iter_rows()]
fields = rows[6]
for row in rows[7:]:
    fields = [f.replace(' ', '_') for f in fields]    
    shot = dict(zip(fields, row))
    shot['lat'] = shot['Latitude-100meter']
    shot['lon'] = shot['Longitude-100meter']
    del shot['Latitude-100meter']
    del shot['Longitude-100meter']
    shot['utm_lat'], shot['utm_lon'], _, _ = utm.from_latlon(
        shot['lat'], shot['lon']
    )    
    ShotSpotterEvent(shot).m.save()